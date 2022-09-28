# IMPORTING IMPORTANT LIBARIES
import openpyxl
import pandas as pd
import requests
from bs4  import BeautifulSoup
import os
from nltk.tokenize import wordpunct_tokenize
import re
import warnings
warnings.filterwarnings('ignore')

# OPENNING THE INPUT FILE
wb = openpyxl.load_workbook('C:\\Users\\p12m9\\Documents\\Python Coding\\Job Assignment\\Job 2 (Web Crawling)\\Input.xlsx')
ws = wb['Sheet1']
df1 = pd.read_excel('C:\\Users\\p12m9\\Documents\\Python Coding\\Job Assignment\\Job 2 (Web Crawling)\\Input.xlsx')

# ACCESSING VALUES OF DIFFERENT COLUMNS
s_no = ws.cell(row=2, column=1).value
s_no = int(s_no) 
i = ws.max_row #getting to know the number of rows in the given table.
e_no = ws.cell(row=i, column=1).value
e_no = int(e_no) +1 #we are adding one so that we can use it for "for loop"
counter = 2 #Initializing counter variable
df1 = pd.read_excel('C:\\Users\\p12m9\\Documents\\Python Coding\\Job Assignment\\Job 2 (Web Crawling)\\Input.xlsx') #creating a dataframe to join it later


# WEB SCRAPING

for j in range(s_no, e_no): #scraping every row in the excel file
    try: #few links were not abe to treat as hyperlink
        link = ws.cell(row=counter, column=2).hyperlink.target
    except: #so we are treating it as value too
        link = ws.cell(row=counter, column=2).value
    headers = {'User-Agent': 'Version 104.0.5112.81 (Windows; Windows 11 64-bit operating system; rv:55.0) Gecko/20100101 Version 104.0.5112.81',} #bot parsing were not allowed
    url = requests.get(link,  headers=headers)
    content = BeautifulSoup(url.content, 'html.parser')
    if content.find('h1') == None: #if the page does not exit
        counter = counter+1
        pass
    else:
        head = content.find('h1').get_text() #getting few text of heading
        if content.find('pre', class_= 'wp-block-preformatted') == None: #if author name is given then let it pass
            pass
        else: #if author name is given then ignoring it
            text = content.find('pre', class_= 'wp-block-preformatted').decompose()
        text = content.find('div', class_= 'td-post-content').get_text() #getting few text of content
        # SAVING THE PARSE CONTENT INTO .txt FILE
        filename = str(ws.cell(row=counter, column=1).value) + '.txt'
        with open(filename, "w", encoding = 'utf-8') as f:
            f.write(str(head) + str(text))
            counter = counter+1
print(df1)


# ANALYSIS

# INITIALIZING EMPTY LIST TO MAKE COLUMN OF DATAFRAME
Z=[]
POSITIVE_SCORE=[]
NEGATIVE_SCORE=[]
POLARITY_SCORE=[]
SUBJECTIVITY_SCORE=[]
AVG_SENTENCE_LENGTH=[]
PERCENTAGE_OF_COMPLEX_WORDS=[]
FOG_INDEX=[]
AVG_NUMBER_OF_WORDS_PER_SENTENCE=[]
COMPLEX_WORD_COUNT=[]
WORD_COUNT=[]
SYLLABLE_PER_WORD=[]
PERSONAL_PRONOUNS=[]
AVG_WORD_LENGTH=[]

# OPENING ONE FILE AT A TIME AND ANALYSIS IT AND STORE IT THE INITIALIZING LIST
Folder_path = 'C:\\Users\\p12m9\\Documents\\Python Coding\\Job Assignment\\Job 2 (Web Crawling)\\Code with Files'
os.chdir(Folder_path)
for Files in os.listdir(): #loading multiple .txt file from the folder one at a time
    if Files.endswith(".txt"):
        z = Files[:-4] #getting filename (removing '.txt') 
        Z.append(z) #adding filename to list
        File_path = f"{Folder_path}\\{Files}"
        with open(File_path, 'r', encoding='utf-8') as f:
            data = f.read()
            for c in ['?','●','.','|','>','<','(',')',',']: #removing few possible puncuation from the file
                if c in data:
                    data=data.replace(c,'')
                    data_list = data.split()
            
            # Word list genration
            path = 'C:\\Users\\p12m9\\Documents\\Python Coding\\Job Assignment\\Job 2 (Web Crawling)\\StopWords'
            os.chdir(path)
            for files in os.listdir(): #loading multiple .txt stopfile from the folder
                if files.endswith(".txt"):
                    file_path = f"{path}\\{files}"
                    with open(file_path, 'r') as f:
                        stop_words = f.read() # Reading Stopwords
                        c = ['?','●','.',',','|','>','<','(',')'] #list of Puncutations
                        stop_word_list = [i for i in stop_words if i not in c] #removing possible Puncutations from text if any
                        word_list = [x.upper() for x in data_list] #making all words in upper case so that all the uppercase stopswords can be identified and removed
                        word_list = [i for i in word_list if i not in stop_words]
                        word_list = [x.lower() for x in word_list] #making all words in lower case so that all the uppercase stopswords can be identified and removed
                        word_list = [i for i in word_list if i not in stop_words] #we have got the list of all the words
            word_count = len(word_list) #we have got the count of the words in a file
            WORD_COUNT.append(word_count) #storing word count into list
			
			# Positive and Negative word count
            with open('C:\\Users\\p12m9\\Documents\\Python Coding\\Job Assignment\\Job 2 (Web Crawling)\\MasterDictionary\\positive-words.txt')as f:
                plist = f.read().splitlines() #spliting all words and passing all words to list
            with open('C:\\Users\\p12m9\\Documents\\Python Coding\\Job Assignment\\Job 2 (Web Crawling)\\MasterDictionary\\negative-words.txt')as f:
                nlist = f.read().splitlines() #spliting all words and passing all words to list
            pscore = 0 #initiallizing pscore which later be increased as we find our positive words
            nscore = 0 #initiallizing nscore which later be increased as we find our negative words
            for w in word_list: #acessing all the word from wordlist
                if w in plist: #if the word is in list of positive word add 1
                    pscore = pscore + 1
                elif w in nlist: #if the word is in list of negative word add 1
                    nscore = nscore + 1
                else:
                    pass
            polarity_score = (pscore - nscore) / ((pscore + nscore) + 0.000001) #calculation of polarity score
            s_score = (pscore + nscore)/ ((len(data_list)) + 0.000001) #calculation of subjectivity score
 
            POSITIVE_SCORE.append(pscore) #storing positive score count into list
            NEGATIVE_SCORE.append(nscore) #storing negaitive score count into list
            POLARITY_SCORE.append(polarity_score) #storing polarity score count into list
            SUBJECTIVITY_SCORE.append(s_score) #storing subjectivity score count into list

            # Sentence list genration
            for c in ['?','\n']:
                if c in data: #accessing data again to take out all the sentence from file
                    data = data.replace(c,'.') #replacing ? or \n(new line) with full stop so that we can easily split
                    sentence_list = data.split('.') #we have got the list of all the sentence from file
                    sentence_list.remove('') #removing blank sentence which might be created my continous full stops if any

            # Calculating syllable per word
            def s(word): #function to count number syllable per word
                count = 0 #initializing the count variable to zero
                vowels = "aeiou" #made a string contating vowels
                if word[0] in vowels: #if the first word starts with vowels then add 1 to count variable
                    count += 1
                for i in range(1, len(word)): #making a range of letters in each words
                    if word[i] in vowels and word[i - 1] not in vowels: #this is to avoids repation ('ee' will be consider as one instead of two)
                        count += 1
                if str(word).endswith("ed"): #if the words endwith "ed" then subtract one
                    count -= 1
                if str(word).endswith("es"): #if the words endwith "ed" then subtract one
                    count -= 1
                if count == 0: #if there is no vowels or sylliable then make count as 1
                    count += 1
                return count
            dic = {} #initializing null dictionary
            key = [] #initializing null list for keys
            value = []  #initializing null list for values
            for word in word_list: #for every word in word list
                k = word #key of dictionary will be the word
                key.append(k)
                v = s(word) #storing function to count number syllable per word in value of our dictionary
                value.append(v)
            syllable_per_word_dic = { key : value for key, value in zip(key, value) } #storing keys and value to dictionary
            total_syllable = sum(syllable_per_word_dic.values()) #adding all syllable
            syllable_per_word = total_syllable / len(word_list) #syllable_per_word = total_syllable / total words
            SYLLABLE_PER_WORD.append(syllable_per_word) #storing syllable per word into list

            # Count of Complex words
            complex_word_list = {i for i in syllable_per_word_dic if syllable_per_word_dic[i]>=int("3")} #acessing dictionary keys and listing it all if count of syllable (value of dictionary) is >= 3
            complex_word_count = 0 #initializing the count variable to zero
            for w in word_list: #for every word in word list
                if w in complex_word_list: #if word matches words in complex words list then add 1
                    complex_word_count  = complex_word_count+1
            COMPLEX_WORD_COUNT.append(complex_word_count) #storing complex word count into list

            avg_sent_len = len(word_list) / len(sentence_list) #Average sentence length = Total word / Total Sentence
            AVG_SENTENCE_LENGTH.append(avg_sent_len) #storing avg setence length into list
            percent_complex = complex_word_count / len(word_list) #Percentage of Complex words = Complex word count / Total number of words
            PERCENTAGE_OF_COMPLEX_WORDS.append(percent_complex) #storing percentage of complex words into list
            Fog_index = 0.4 * (avg_sent_len + percent_complex) #Fog Index as per formula
            FOG_INDEX.append(Fog_index) #storing fox index value into list

            # Average words per sentence
            m = [] #initializing null list 
            for s in sentence_list: #for every sentence in sentence list 
                word_count_per_sentence = 0 #Initializing the count variable to zero
                word_per_sentence = wordpunct_tokenize(s) #Tokenizing words in a sentence with Puncutations sepearted as well
                c = ['?','●','.',',','|','>','<','(',')'] #list of puncutations
                word_per_sentence = [i for i in word_per_sentence if i not in c] #removing puncutations
                for w in word_per_sentence: #for all the words in sentence
                    if w in word_list: #if word is in word list add 1 to count variable
                        word_count_per_sentence = word_count_per_sentence+1
                m.append(word_count_per_sentence) #making a list of all words in every sentence
            Avg_words_per_sent = sum(m) / len(m) #taking out average
            AVG_NUMBER_OF_WORDS_PER_SENTENCE.append(Avg_words_per_sent) #storing avg words per sentence into list

            # Average word length
            c = ['?','●','.',',','|','>','<','(',')',' '] #list of characters
            char = [i for i in data if i not in c] #removing List of Characters from data
            Avg_word_length = len(char) / len(word_list) #Average word length = Total char / total words
            AVG_WORD_LENGTH.append(Avg_word_length) #storing avg words length into list

            # Personal pronouns
            Personal_pronouns = re.findall('i|I|WE|We|we|MY|My|my|OURS|Ours|ours|OUR|Our|our|us',data) #finding list of all possible pronouns in data
            Personal_pronouns_count = len(Personal_pronouns)
            PERSONAL_PRONOUNS.append(Personal_pronouns_count) #storing personal pronoun count into list

# SAVING ALL LIST INTO DATAFRAME AS COLUMNS
df2= pd.DataFrame({'URL_ID':Z,'POSITIVE_SCORE': POSITIVE_SCORE,'NEGATIVE_SCORE': NEGATIVE_SCORE,
    'POLARITY_SCORE': POLARITY_SCORE,'SUBJECTIVITY_SCORE': SUBJECTIVITY_SCORE,'AVG_SENTENCE_LENGTH': AVG_SENTENCE_LENGTH,
    'PERCENTAGE_OF_COMPLEX_WORDS': PERCENTAGE_OF_COMPLEX_WORDS,'FOG_INDEX': FOG_INDEX,
    'AVG_NUMBER_OF_WORDS_PER_SENTENCE': AVG_NUMBER_OF_WORDS_PER_SENTENCE,'COMPLEX_WORD_COUNT': COMPLEX_WORD_COUNT,
    'WORD_COUNT':WORD_COUNT,'SYLLABLE_PER_WORD':SYLLABLE_PER_WORD,'PERSONAL_PRONOUNS':PERSONAL_PRONOUNS,
    'AVG_WORD_LENGTH': AVG_WORD_LENGTH})
df2["URL_ID"] = df2["URL_ID"].astype(float)
print(df2)

# JOINING INPUT DATAFRAME & ANALITICAL DATAFRAME AND SAVING INTO CSV
df = pd.merge(df1,df2, on = 'URL_ID', how = 'outer')
df.to_csv(r'C:\\Users\\p12m9\\Documents\\Python Coding\\Job Assignment\\Job 2 (Web Crawling)\\Output Data Structure.csv', index=False)
print(df)